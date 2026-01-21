"""
Export Service - Multi-format data export
Supports: CSV, Excel, JSON, PDF, GeoJSON
"""

import csv
import json
import io
from datetime import datetime
from typing import List, Dict, Any, Optional

# Excel export
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PDF export
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image
from reportlab.lib.enums import TA_CENTER, TA_LEFT

# GeoJSON
import geojson


class ExportService:
    """
    Service for exporting data in multiple formats
    """
    
    SUPPORTED_FORMATS = ['csv', 'xlsx', 'json', 'pdf', 'geojson']
    
    @staticmethod
    def export(data: List[Dict], columns: List[Dict], format: str, 
               title: str = "TEDI Export", sector: str = "") -> tuple:
        """
        Export data in the specified format
        
        Args:
            data: List of dictionaries containing the data
            columns: List of column definitions [{'key': 'field', 'label': 'Label'}]
            format: Export format (csv, xlsx, json, pdf, geojson)
            title: Title for the export
            sector: Sector name for context
            
        Returns:
            (content: bytes, content_type: str, filename: str)
        """
        format = format.lower()
        
        if format not in ExportService.SUPPORTED_FORMATS:
            raise ValueError(f"Unsupported format: {format}. Supported: {ExportService.SUPPORTED_FORMATS}")
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_filename = f"tedi_{sector}_{timestamp}" if sector else f"tedi_export_{timestamp}"
        
        if format == 'csv':
            return ExportService._export_csv(data, columns, base_filename)
        elif format == 'xlsx':
            return ExportService._export_excel(data, columns, title, base_filename)
        elif format == 'json':
            return ExportService._export_json(data, columns, title, sector, base_filename)
        elif format == 'pdf':
            return ExportService._export_pdf(data, columns, title, sector, base_filename)
        elif format == 'geojson':
            return ExportService._export_geojson(data, columns, sector, base_filename)
    
    @staticmethod
    def _export_csv(data: List[Dict], columns: List[Dict], filename: str) -> tuple:
        """Export to CSV format"""
        output = io.StringIO()
        
        # Write headers
        headers = [col['label'] for col in columns]
        writer = csv.writer(output)
        writer.writerow(headers)
        
        # Write data
        for row in data:
            row_data = []
            for col in columns:
                value = ExportService._get_nested_value(row, col['key'])
                row_data.append(ExportService._format_value(value))
            writer.writerow(row_data)
        
        content = output.getvalue().encode('utf-8-sig')  # BOM for Excel compatibility
        return content, 'text/csv; charset=utf-8', f"{filename}.csv"
    
    @staticmethod
    def _export_excel(data: List[Dict], columns: List[Dict], 
                      title: str, filename: str) -> tuple:
        """Export to Excel format with formatting"""
        wb = Workbook()
        ws = wb.active
        ws.title = "Data"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A5F7A", end_color="1A5F7A", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title row
        ws.merge_cells('A1:' + get_column_letter(len(columns)) + '1')
        ws['A1'] = title
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal="center")
        
        # Date row
        ws.merge_cells('A2:' + get_column_letter(len(columns)) + '2')
        ws['A2'] = f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws['A2'].alignment = Alignment(horizontal="center")
        ws['A2'].font = Font(italic=True, size=10)
        
        # Headers (row 4)
        for col_idx, col in enumerate(columns, 1):
            cell = ws.cell(row=4, column=col_idx, value=col['label'])
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        
        # Data rows
        for row_idx, row in enumerate(data, 5):
            for col_idx, col in enumerate(columns, 1):
                value = ExportService._get_nested_value(row, col['key'])
                cell = ws.cell(row=row_idx, column=col_idx, value=ExportService._format_value(value))
                cell.border = thin_border
                
                # Align numbers to right
                if isinstance(value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
        
        # Adjust column widths
        for col_idx, col in enumerate(columns, 1):
            max_length = len(col['label'])
            for row_idx in range(5, min(len(data) + 5, 100)):  # Check first 100 rows
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
        
        # Save to bytes
        output = io.BytesIO()
        wb.save(output)
        content = output.getvalue()
        
        return content, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', f"{filename}.xlsx"
    
    @staticmethod
    def _export_json(data: List[Dict], columns: List[Dict], 
                     title: str, sector: str, filename: str) -> tuple:
        """Export to JSON format with metadata"""
        export_data = {
            "metadata": {
                "title": title,
                "sector": sector,
                "exported_at": datetime.now().isoformat(),
                "total_records": len(data),
                "columns": [{"key": col['key'], "label": col['label']} for col in columns]
            },
            "data": data
        }
        
        content = json.dumps(export_data, indent=2, ensure_ascii=False, default=str).encode('utf-8')
        return content, 'application/json; charset=utf-8', f"{filename}.json"
    
    @staticmethod
    def _export_pdf(data: List[Dict], columns: List[Dict], 
                    title: str, sector: str, filename: str) -> tuple:
        """Export to PDF format with table"""
        output = io.BytesIO()
        
        # Use landscape for more columns
        doc = SimpleDocTemplate(
            output, 
            pagesize=landscape(A4),
            leftMargin=1*cm,
            rightMargin=1*cm,
            topMargin=1*cm,
            bottomMargin=1*cm
        )
        
        elements = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            alignment=TA_CENTER,
            spaceAfter=20
        )
        elements.append(Paragraph(title, title_style))
        
        # Subtitle with date
        subtitle_style = ParagraphStyle(
            'Subtitle',
            parent=styles['Normal'],
            fontSize=10,
            alignment=TA_CENTER,
            textColor=colors.gray,
            spaceAfter=20
        )
        elements.append(Paragraph(
            f"Sector: {sector.upper()} | Exported: {datetime.now().strftime('%Y-%m-%d %H:%M')} | Records: {len(data)}",
            subtitle_style
        ))
        
        # Limit data for PDF (too many rows = too many pages)
        max_rows = 100
        display_data = data[:max_rows]
        if len(data) > max_rows:
            elements.append(Paragraph(
                f"<i>Note: Showing first {max_rows} of {len(data)} records. Use CSV/Excel for full data.</i>",
                styles['Normal']
            ))
            elements.append(Spacer(1, 10))
        
        # Build table data
        # Limit columns for readability
        display_columns = columns[:8]  # Max 8 columns for PDF
        
        table_data = []
        # Headers
        headers = [col['label'][:15] for col in display_columns]  # Truncate long headers
        table_data.append(headers)
        
        # Data rows
        for row in display_data:
            row_data = []
            for col in display_columns:
                value = ExportService._get_nested_value(row, col['key'])
                formatted = str(ExportService._format_value(value))[:20]  # Truncate
                row_data.append(formatted)
            table_data.append(row_data)
        
        # Create table
        col_widths = [2*cm] * len(display_columns)
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Table style
        style = TableStyle([
            # Header
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1A5F7A')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data rows
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            
            # Alternating row colors
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F5F5F5')]),
            
            # Grid
            ('GRID', (0, 0), (-1, -1), 0.5, colors.gray),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        table.setStyle(style)
        elements.append(table)
        
        # Footer
        elements.append(Spacer(1, 20))
        footer_style = ParagraphStyle(
            'Footer',
            parent=styles['Normal'],
            fontSize=8,
            alignment=TA_CENTER,
            textColor=colors.gray
        )
        elements.append(Paragraph(
            "TEDI - Territorial & Economic Data Index | tedi.africa",
            footer_style
        ))
        
        doc.build(elements)
        content = output.getvalue()
        
        return content, 'application/pdf', f"{filename}.pdf"
    
    @staticmethod
    def _export_geojson(data: List[Dict], columns: List[Dict], 
                        sector: str, filename: str) -> tuple:
        """Export to GeoJSON format (for geographic data)"""
        features = []
        
        for row in data:
            # Try to get coordinates
            lat = row.get('latitude') or row.get('lat')
            lon = row.get('longitude') or row.get('lon') or row.get('lng')
            
            # Try to get from nested commune object
            if not lat or not lon:
                commune = row.get('commune', {})
                if isinstance(commune, dict):
                    lat = commune.get('latitude') or commune.get('center_lat')
                    lon = commune.get('longitude') or commune.get('center_lon')
            
            # Create properties from all columns
            properties = {}
            for col in columns:
                value = ExportService._get_nested_value(row, col['key'])
                properties[col['label']] = ExportService._format_value(value)
            
            # Add original ID if available
            if 'id' in row:
                properties['id'] = row['id']
            
            # Create feature
            if lat and lon:
                try:
                    point = geojson.Point((float(lon), float(lat)))
                    feature = geojson.Feature(
                        geometry=point,
                        properties=properties
                    )
                    features.append(feature)
                except (ValueError, TypeError):
                    # If coords invalid, add without geometry
                    feature = geojson.Feature(
                        geometry=None,
                        properties=properties
                    )
                    features.append(feature)
            else:
                # No coordinates, add without geometry
                feature = geojson.Feature(
                    geometry=None,
                    properties=properties
                )
                features.append(feature)
        
        feature_collection = geojson.FeatureCollection(features)
        
        # Add metadata
        feature_collection['metadata'] = {
            'sector': sector,
            'exported_at': datetime.now().isoformat(),
            'total_features': len(features)
        }
        
        content = geojson.dumps(feature_collection, indent=2).encode('utf-8')
        return content, 'application/geo+json; charset=utf-8', f"{filename}.geojson"
    
    @staticmethod
    def _get_nested_value(data: Dict, key: str) -> Any:
        """Get value from nested dict using dot notation or direct key"""
        if '.' in key:
            parts = key.split('.')
            value = data
            for part in parts:
                if isinstance(value, dict):
                    value = value.get(part)
                else:
                    return None
            return value
        
        # Handle special cases like 'commune' -> 'commune.name'
        value = data.get(key)
        if isinstance(value, dict) and 'name' in value:
            return value['name']
        return value
    
    @staticmethod
    def _format_value(value: Any) -> Any:
        """Format value for export"""
        if value is None:
            return ''
        if isinstance(value, bool):
            return 'Yes' if value else 'No'
        if isinstance(value, float):
            # Round to reasonable precision
            if abs(value) < 0.01:
                return round(value, 6)
            elif abs(value) < 1:
                return round(value, 4)
            elif abs(value) < 1000:
                return round(value, 2)
            else:
                return round(value, 0)
        if isinstance(value, dict):
            return value.get('name', str(value))
        if isinstance(value, list):
            return ', '.join(str(v) for v in value)
        return value
